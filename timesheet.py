import openpyxl
from openpyxl import load_workbook
import os
import glob
import re
import locale
from openpyxl.comments import Comment
import json
import pdb
import sys

locale.setlocale(locale.LC_ALL, 'sv_SE.utf8') # sv_SE in linux, 'sv' in win

# Fixed rows for work/leavehours
dateRow = 1
freeRow = 2
vabRow = 3
illRow = 4
holiRow = 5
workRow = 6
projRow = 7


def toDecTime(hrs, mins):
    timeVal = int(hrs) + int(mins) / 60.0
    return round(timeVal, 2)

def autosize_columns(worksheet):
    for col in worksheet.iter_cols(min_row=0,max_row=0,max_col=36):
        max_length = 0
        column = col[0].column_letter # Get the column name
        # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
        adjusted_width = 0
        for c in col:
            try: # Necessary to avoid error on empty cells
                if c.row == workRow and c.col_idx > 3:  # Hide columns with zero worked hours
                    if c.value == 0.0:
                        worksheet.column_dimensions[column].hidden = True
                        break
                if len(str(c.value)) > max_length:
                    max_length = len(c.value)
            except:
                pass
        adjusted_width = (max_length + 2)

        worksheet.column_dimensions[column].width = adjusted_width

def getProjNrs():
    return list( json.loads(open('projs.json','r').read()).keys() )

def getProjCmts():
    return list( json.loads(open('projs.json','r').read()).values() )



# Column will be incremented with every new date we find in the input.
dateCol = 3  # First columns saved for project number and sum

book = openpyxl.Workbook()
sheet1 = book.worksheets[0]

# Populate first column with titles
sheet1.cell(row=freeRow, column=2).value = 'F-ledig'
sheet1.cell(row=illRow, column=2).value = 'Sjuk'
sheet1.cell(row=vabRow, column=2).value = 'VAB'
sheet1.cell(row=holiRow, column=2).value = 'Semester'
sheet1.cell(row=workRow, column=2).value = 'Jobb'
sheet1.cell(row=projRow, column=1).value = 'Name'
sheet1.cell(row=projRow, column=2).value = 'ProjNr'
sheet1.cell(row=projRow, column=3).value = 'Sum[h]:'

# List of my current project numbers
projNrs = getProjNrs()
comment = getProjCmts()
print("This month's projects:")
# Fill in projects and sum up the hours per row
for row in range(0, len(projNrs)):
    rowNr = projRow+row+1

    cmmtCell = sheet1.cell(row=rowNr, column=1)
    cmmtCell.value = comment[row]
    projCell = sheet1.cell(row=rowNr, column=2)
    projCell.value = projNrs[row]
    print('{0} : {1} '.format(projNrs[row], comment[row]))
    sumCell = sheet1.cell(row=rowNr, column=3)
    sumCell.value = '=sum($D${0}:$AK${0})'.format(rowNr)

sheet1.cell(row=workRow, column=3).value = '=sum(D{0}:AK{0})-SUM(C{1}:C{2})'.format(workRow,projRow+1,projRow+len(projNrs))

sheet1.auto_filter.ref = 'A' + str(projRow) + ':C' + str(projRow + len(projNrs))


# Get all inputfiles from onedrive
inpGlob = sys.argv[1] + 'Saldon*.xlsx'
print('Reading saldo from \"' + inpGlob + '\"')
inpFiles = glob.glob(inpGlob)

for file in inpFiles:
    print("Reading " + file + '...')
    wb = load_workbook(file, read_only=True)
    print('done')
    inpSheet = wb.worksheets[0]
    cRow = 0
    for row in inpSheet.iter_rows():
        cRow += 1
        if cRow < 7:  # Skip first header rows
            print('Skipping headers')
            next
        line = ''
        for cell in row:
            if cell.value:
                line += str(cell.value)
                line += ' '
        # print(line)
        m = re.search('^(20\d\d-\d\d-\d\d)', line)
        if m:  # Found a new date, increment column and note the date
            # print(line)
            dateCol += 1
            sheet1.cell(row=dateRow, column=dateCol).value = '"' + m.group(1) + '"'
        m = re.search('Tillf\.för\.ledig\s*(\d*):(\d*)', line)
        if m:
            # print(line)
            sheet1.cell(row=vabRow, column=dateCol).value = toDecTime(m.group(1), m.group(2))
        m = re.search('Föräldraledig\s*(\d*):(\d*)', line)
        if m:
            sheet1.cell(row=freeRow, column=dateCol).value = toDecTime(m.group(1), m.group(2))
        m = re.search('Närvarotid\s*(\d*):(\d*)', line)
        if m:
            # print(line)
            sheet1.cell(row=workRow, column=dateCol).value = toDecTime(m.group(1), m.group(2))
        m = re.search('Semester\s*(\d*):(\d*)', line)
        if m:
            # print(line)
            sheet1.cell(row=holiRow, column=dateCol).value = toDecTime(m.group(1), m.group(2))
        m = re.search('Sjukdom\s*(\d*):(\d*)', line)
        if m:
            # print(line)
            sheet1.cell(row=illRow, column=dateCol).value = toDecTime(m.group(1), m.group(2))

autosize_columns(sheet1)

print("Parsing done!")

book.save('timesheet.xlsx')
